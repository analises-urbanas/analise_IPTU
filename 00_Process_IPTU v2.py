## 00_Process_IPTU: o programa é um processador dos arquivos do IPTU, permitindo automatizar...
#...o preparo das bases para o PosgreSQL e XLSX, dos diversos anos do IPTU.

'''
INICIO DE DESENV: 10/2018
POR: Marcelo Baliu Fiamenghi
ma.baliu@gmail.com
(11) 9.8647.7960
---------------------------
* * DICIONÁRIO * *

**
'''

import zipfile, os, psycopg2, subprocess, dir_path
import sql_script as s


# Criação do diretório onde se encontra os arquivos dos IPTU
folder = dir_path.folders()
# Criação da lista dos IPTU presentes na pasta
files = os.listdir(folder)
psql = '"C:\\Program Files (x86)\\PostgreSQL\\10\\bin\\psql"'


def env():
  global folder, files, psql

  # Criação do diretório onde se encontra os arquivos dos IPTU
  folder = dir_path.folders()
  # Criação da lista dos IPTU presentes na pasta
  files = os.listdir(folder)
  psql = '"C:\\Program Files (x86)\\PostgreSQL\\10\\bin\\psql"'


def p(Funcao):
  print('Executando {1}'.format(Funcao))


def cmd_exec(cmd_line):
  # Executa uma string no CMD LINE

  process = subprocess.Popen(cmd_line, stdout=subprocess.PIPE)
  #print(cmd_line)
  for line in iter(process.stdout.readline, b''):
    print(line)


# FOR LOOP -----
def unzip_file(File):
  global file_dir, fObj
  env()

  file_dir = os.path.join(folder,File)
  # print(f)
  # print(file_dir)

  # ACESSO AO ZIPFILE
  with zipfile.ZipFile(file_dir) as zObj:
    # NOME DO ARQUIVO: fileObject
    fObj = zObj.namelist()

    # EXTRAÇÃO
    zObj.extractall (folder)

def find_csv():
  # Obtém o nome do arquivo e o ano de execução

  global csv_file, ano_exercicio
  env()
  files_zip = os.listdir(folder)

  for i in files_zip:
    if i.endswith('.csv'):
      csv_file = i

  ano_exercicio = csv_file[:-4][-4:]



def new_schema(Ano):
  # Cria um novo schema

  env()
  # find_csv()

  nome_schema = 'a' + str(Ano)


  # nw_schema: new schema - versão psql
  cmd_line = psql + ' -U postgres -d iptu_teste -c "create schema if not exists {0}"'.format(nome_schema)
  process = subprocess.Popen(cmd_line, stdout=subprocess.PIPE)
  #print(cmd_line)
  for line in iter(process.stdout.readline, b''):
    print(line)

  # new_table(ano_exercicio)
  # copy_imp(ano_exercicio)
  # END

def new_table(Ano):
  # Cria uma tabela completa do ano de exercício

  env()
  import sql_script as s
  nome_schema = 'a' + Ano
  nome_tabela = '_' + Ano

  # nw_table: new table - versão psql
  # s.creatbl(var1, var2) = CREATE TABLE sch.table (..)
  cmd_line = psql + ' -U postgres -d iptu_teste -c "{0}" '.format(s.create_tbl(nome_schema, nome_tabela))
  process = subprocess.Popen(cmd_line, stdout=subprocess.PIPE)
  for line in iter(process.stdout.readline, b''):
    print(line)

  # print (cmd_line)

def copy_imp(Ano):
  # COPY_IMPort
  # IMPORTA o csv bruto do IPTU para o postgres, com o Schema: a_2001 e Tabela: _2001, sendo a mais bruta
  env()
  find_csv()

  file_dir = os.path.join(folder, csv_file)
  nome_schema = 'a' + Ano
  nome_tabela = '_' + Ano

  #COPY <dbname> [(col1, col2, col3)] FROM ‘.\dir’ DELIMITER ‘,’ CSV;
  # SET client_encoding TO WIN1252;
  qry = "SET client_encoding TO WIN1252;\
  COPY {0}.{1} FROM {2} DELIMITER ';' CSV HEADER;".format(nome_schema, nome_tabela, file_dir)
  #COPY {0}.{1} [(<colunas>)] FROM 'file_dir' DELIMITER ',' CSV HEADER".format(nome_schema, nome_tabela)

  #COPY db.iptu_centro (id, "NUMERO DO CONTRIBUINTE", "ANO DO EXERCICIO", "NUMERO DA NL", "DATA DO CADASTRAMENTO", "TIPO DE CONTRIBUINTE 1", "CPF/CNPJ DO CONTRIBUINTE 1", "NOME DO CONTRIBUINTE 1", "TIPO DE CONTRIBUINTE 2", "CPF/CNPJ DO CONTRIBUINTE 2", "NOME DO CONTRIBUINTE 2", "NUMERO DO CONDOMINIO", "CODLOG DO IMOVEL", "NOME DE LOGRADOURO DO IMOVEL", "NUMERO DO IMOVEL", "COMPLEMENTO DO IMOVEL", "BAIRRO DO IMOVEL", "REFERENCIA DO IMOVEL", "CEP DO IMOVEL", "QUANTIDADE DE ESQUINAS/FRENTES", "FRACAO IDEAL", "AREA DO TERRENO", "AREA CONSTRUIDA", "AREA OCUPADA", "VALOR DO M2 DO TERRENO", "VALOR DO M2 DE CONSTRUCAO", "ANO DA CONSTRUCAO CORRIGIDO", "QUANTIDADE DE PAVIMENTOS", "TESTADA PARA CALCULO", "TIPO DE USO DO IMOVEL", "TIPO DE PADRAO DA CONSTRUCAO", "TIPO DE TERRENO", "FATOR DE OBSOLESCENCIA", "ANO DE INICIO DA VIDA DO CONTRIBUINTE", "MES DE INICIO DA VIDA DO CONTRIBUINTE", "FASE DO CONTRIBUINTE") FROM stdin;

  # Reconhecer se a tabela ja tem DADOS. Evitar duplicar os dados ===============================

  cmd_line = psql + ' -U postgres -d iptu_teste -c "{0}" '.format(qry)
  process = subprocess.Popen(cmd_line, stdout=subprocess.PIPE)
  for line in iter(process.stdout.readline, b''):
    print(line)

  # print(qry)

def rmvCSV(File):
  # rmvCSV: remove CSV
  # Deleta o arquivo csv já copiado
  os.unlink(os.path.join(folder, File))
  

def sqlc(Ano):
  # Adiciona o campo SQLC (SQL Condominial mais o condomínio)
  import sql_script as s

  nome_schema = 'a' + str(Ano)
  nome_tabela = '_' + str(Ano)

  '''
  ------------------------------------------------------------------------------------------------
  # Teste inicial de criação do script para criação do SQLC
  qry = 'ALTER TABLE {0}.{1} ADD COLUMN \"SQLC\" varchar;\
  UPDATE {0}.{1}\
  SET \"SQLC\" = \
  WHERE '

  
  (CASE 
    WHEN left(iptu16_centro."NUMERO DO CONDOMINIO",2)  <> '00' THEN concat(left(iptu16_centro."NUMERO DO CONTRIBUINTE",6), '0000', left(iptu16_centro."NUMERO DO CONDOMINIO",2))
    WHEN left(iptu16_centro."NUMERO DO CONDOMINIO",2)  = '00' THEN concat(left(iptu16_centro."NUMERO DO CONTRIBUINTE",10), left(iptu16_centro."NUMERO DO CONDOMINIO",2))
  END)
  qry = qry.format(nome_schema, nome_tabela)
  ------------------------------------------------------------------------------------------------
  '''



  qry = s.create_sqlc_values(nome_schema, nome_tabela)

  cmd_line = psql + ' -U postgres -d iptu_teste -c "{0}" '.format(qry)
  # Verificação do cmd
  # print(cmd_line)

  process = subprocess.Popen(cmd_line, stdout=subprocess.PIPE)
  for line in iter(process.stdout.readline, b''):
    print(line)

def exec_sqlc(Ano):
  # Executa o sqlc(Ano) contabilizando o tempo

  # =============================================================
  import time
  time_ini = time.time()
  print('==== Executando sqlc() ....')

  sqlc(Ano)

  print('==== sqlc() concluido!')
  time_fim = time.time()

  duracao = time_fim - time_ini
  duracao_display = str(int(duracao//60)) + ':' + str(int(duracao%60))
  print('==== Duração: ' + duracao_display + ' minutos.\n')


def crt_views(Ano):
  # Cria as views para cada região
  # Cria as VIEWs para as 8 regiões de São Paulo escolhidas, por meio de seus setores

  import sql_script as s
  nome_schema = 'a' + str(Ano)
  nome_tabela = '_' + str(Ano)


  qry = s.create_views(nome_schema, nome_tabela)
  print(qry)
  cmd_line = psql + ' -U postgres -d iptu_teste -c "{0}" '.format(qry)
  print(cmd_line)
  cmd_exec(cmd_line)


def pg2excel(Ano):
  # Exporta cada view para um XLSX

  nome_schema = 'a' + str(Ano)
  nome_tabela = '_' + str(Ano)
  # nome_view = ['centro','oeste','leste_1','leste_2','norte_1','norte_2','sul_1','sul_2']
  nome_view = ['sul_2']


  for v in nome_view:
    xls_name = 'iptu_{0}_{1}.xlsx'.format(str(Ano), v)
    cmd_line = 'py pg2xls.py -U postgres -d db_name -t {0}.{1} -f {3} -T iptu{2}'
    cmd_line = cmd_line.format(nome_schema, v, str(ano), xls_name)

    print(cmd_line)
    aval(cmd_exec, cmd_line)
    # cmd_exec(cmd_line)

def filterExcel(Ano):
  # Cria a aba de filtro no novo excel criado


  import openpyxl, os
  from openpyxl.cell.cell import get_column_letter
  from openpyxl.styles import PatternFill, Border, Alignment, Side, Font

  # Cria o caminho para os arquivos XLS
  folder = r'C:\_RAW\TESTE'
  file_name = 'table.xlsx'
  file_path = os.path.join(folder, file_name)

  # Conecta ao XLS e cria a planilha FILTRO
  wb = openpyxl.load_workbook(file_path)
  ws = wb.create_sheet('FILTRO')

#ESTILO DA TABELA ORIGINAL


  sheet = wb['FILTRO']

  # CONFIGURA O CABEÇALHO DO FILTRO
  #   TEXTO
  header = ['Sqlc','Numero do contribuinte','Ano do exercicio','Numero da nl','Data do cadastramento','Tipo de contribuinte 1','Cpf/cnpj do contribuinte 1','Nome do contribuinte 1','Tipo de contribuinte 2','Cpf/cnpj do contribuinte 2','Nome do contribuinte 2','Numero do condominio','Codlog do imovel','Nome de logradouro do imovel','Numero do imovel','Complemento do imovel','Bairro do imovel','Referencia do imovel','Cep do imovel','Quantidade de esquinas/frentes','Fracao ideal','Area do terreno','Area construida','Area ocupada','Valor do m2 do terreno','Valor do m2 de construcao','Ano da construcao corrigido','Quantidade de pavimentos','Testada para calculo','Tipo de uso do imovel','Tipo de padrao da construcao','Tipo de terreno','Fator de obsolescencia','Ano de inicio da vida do contribuinte','Mes de inicio da vida do contribuinte','Fase do contribuinte']
  #   ESTILO
  thin = Side(border_style="thin", color="ffffff")
  medium = Side(border_style="medium", color="000000")
  border = Border(top=thin, left=thin, right=thin, bottom=medium)
  fill = PatternFill("solid", fgColor="048A7D")
  al = Alignment(horizontal="left", vertical="center")
  font = Font(color="ffffff")

  sheet.row_dimensions[1].height = 20

  for h in range(len(header)):
      c = sheet.cell(row=1, column=h+1)
      letter = openpyxl.cell.cell.get_column_letter(h+1)
      c.value = header[h]
      c.fill = fill
      c.alignment = al
      c.border = border
      c.font = font
      sheet.column_dimensions[letter].width = 23


  file_name2 = 'table_cor.xlsx'
  file_path2 = os.path.join(folder, file_name2)
  wb.save(file_path2)




############# IPTU Contribuintes ##################
  
def create_iptuDB(Ano):
  # create_iptuDB(): processo individual - Execução linear e individual das funções para um ano de execução

  import time
  env()
  #retornará uma lista de arquivos

  # ano = input('Qual o ano do IPTU?\n')
  arquivo = 'IPTU_' + str(Ano) + '.zip'

  p('unzip_file')
  unzip_file(arquivo)

  p('find_csv')
  find_csv()
  new_schema(Ano)
  new_table(Ano)
  p('copy_imp')
  copy_imp(Ano)
  # Deleta o csv antigo

  time.sleep(5)
  rmvCSV(csv_file)

  p('exec_sqlc')
  exec_sqlc(Ano)
  crt_views(Ano)
  p('pg2excel')
  pg2excel(Ano)
  filterExcel(Ano)
  print('Concluído')


def update_iptuDB(Ano):
  env()


def multi_process(Ano_inicio, Ano_fim):
  Ano_inicio = int(Ano_inicio)
  Ano_fim = int(Ano_fim)

  for i in range(Ano_inicio, Ano_fim+1):
    print('Processando o IPTU do ano {1}'.format(i))

    create_iptuDB(i)
  



# ============= BACK UP =========================
# BACKUP: Exposta (em formato binário) o Schema por ano, contendo as tabelas e as vistas (views)
# BACKUP: Compacta os arquivos binários criados


# ====================== MELHORAR ===============================
# Atualizar o tipo de dados dos campos
# Atualizador das tabelas e dos Excel



# ====================== SQLC =====================================
# Agregar dados segundo o dado do SQLC, para obtenção dos dados territoriais
# Cria a tabela do IPTU agregada por território
# ... repete as funções





### Correção dos nomes dos IPTU

def aval(Funcao, Dado):
  def time_aval():
    import time
    time_ini = time.time()
    print('==== Executando....')

    Funcao(Dado)

    print('==== concluido!')
    time_fim = time.time()

    duracao = time_fim - time_ini
    duracao_display = str(int(duracao//60)) + ':' + str(int(duracao%60))
    print('==== Duração: ' + duracao_display + ' minutos.\n')
  return time_aval
